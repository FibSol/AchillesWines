"""
WERC Global Wine Markets megafile importer.

Source: https://economics.adelaide.edu.au/wine-economics/databases
File:   megafile_of_global_wine_data_1835_to_2024-0425.xlsx  (~17 MB)
Auth:   None (plain HTTP GET)
Cadence: Annual (file updated ~December each year)

Imports two sheets from the 111-sheet megafile:
  T1 → vine_area_kha    (thousands of hectares of vineyard)
  T6 → wine_production_kl (kiloliters of wine produced)

Layout (consistent across all sheets):
  Row 1  : Title string with unit in parentheses
  Row 2  : Country names across columns (col A is empty)
  Row 3+ : Year (col A) + values per country

Both metrics feed into fact_werc_stats for vintage quality context and
historical trend analysis. Only wine-producing countries are stored; regional
aggregates ("Other WEM", "World", etc.) are skipped.
"""
import io
import os
import sqlite3
import tempfile
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import httpx
from rich.console import Console

from .base import BaseScraper, ScrapeResult
from ..dlq import write_dlq

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

console = Console()

_MEGAFILE_URL = (
    "https://economics.adelaide.edu.au/wine-economics/ua/media/438/"
    "megafile_of_global_wine_data_1835_to_2024-0425.xlsx"
)

# Sheets to import: (sheet_name_fragment, metric_code, unit)
_SHEETS = [
    ("T1 Vine area",        "vine_area_kha",      "'000 ha"),
    ("T6 Wine production",  "wine_production_kl",  "KL"),
]

# Full English name → ISO-3166-1 alpha-2
_COUNTRY_MAP: dict[str, str] = {
    "France": "FR",
    "Italy": "IT",
    "Portugal": "PT",
    "Spain": "ES",
    "Austria": "AT",
    "Belgium": "BE",
    "Bel-Lux": "BE",
    "Denmark": "DK",
    "Finland": "FI",
    "Germany": "DE",
    "Greece": "GR",
    "Ireland": "IE",
    "Netherlands": "NL",
    "Sweden": "SE",
    "Switzerland": "CH",
    "United Kingdom": "GB",
    "Bulgaria": "BG",
    "Croatia": "HR",
    "Georgia": "GE",
    "Hungary": "HU",
    "Moldova": "MD",
    "Romania": "RO",
    "Russia": "RU",
    "Ukraine": "UA",
    "Australia": "AU",
    "New Zealand": "NZ",
    "Canada": "CA",
    "United States": "US",
    "Argentina": "AR",
    "Brazil": "BR",
    "Chile": "CL",
    "Mexico": "MX",
    "Uruguay": "UY",
    "Algeria": "DZ",
    "Morocco": "MA",
    "South Africa": "ZA",
    "Tunisia": "TN",
    "Turkey": "TR",
    "China": "CN",
    "Hong Kong": "HK",
    "India": "IN",
    "Japan": "JP",
    "Korea": "KR",
    "Luxembourg": "LU",
    "Czech Republic": "CZ",
    "Slovakia": "SK",
    "Slovenia": "SI",
    "Serbia": "RS",
    "Montenegro": "ME",
    "North Macedonia": "MK",
    "Albania": "AL",
    "Bosnia": "BA",
}

_ENSURE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS fact_werc_stats (
    stat_id       integer PRIMARY KEY AUTOINCREMENT NOT NULL,
    source_key    integer NOT NULL REFERENCES dim_source(source_key),
    country_code  text    NOT NULL,
    year          integer NOT NULL,
    metric        text    NOT NULL,
    value         real    NOT NULL,
    unit          text    NOT NULL,
    batch_id      text    NOT NULL,
    created_at    integer NOT NULL DEFAULT (unixepoch())
);
"""
_ENSURE_IDX_UNIQUE   = "CREATE UNIQUE INDEX IF NOT EXISTS idx_werc_unique ON fact_werc_stats(country_code, year, metric);"
_ENSURE_IDX_COUNTRY  = "CREATE INDEX IF NOT EXISTS idx_werc_country_year ON fact_werc_stats(country_code, year);"
_ENSURE_IDX_METRIC   = "CREATE INDEX IF NOT EXISTS idx_werc_metric ON fact_werc_stats(metric);"


def _get_source_key(conn: sqlite3.Connection, source_code: str) -> int:
    row = conn.execute(
        "SELECT source_key FROM dim_source WHERE source_code = ?", (source_code,)
    ).fetchone()
    if row is None:
        raise RuntimeError(
            f"dim_source row missing for '{source_code}'. Run migration 0009."
        )
    return row[0]


def _find_sheet(wb, name_fragment: str):
    """Return the worksheet whose name contains name_fragment (case-insensitive)."""
    frag = name_fragment.lower()
    for name in wb.sheetnames:
        if frag in name.lower():
            return wb[name]
    return None


def _parse_sheet(ws, metric: str, unit: str) -> list[tuple[str, int, float]]:
    """
    Parse a WERC data sheet.

    Layout:
      Row 1: title (ignored)
      Row 2: [None, CountryA, CountryB, ...]
      Row 3+: [year, valueA, valueB, ...]

    Returns list of (country_code, year, value) tuples, skipping:
      - Countries not in _COUNTRY_MAP (aggregates, "Other X", "World")
      - Rows where year is not an integer
      - None values
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    # Row index 1 (0-based) = row 2 = country headers
    headers = rows[1]   # (None, "France", "Italy", ...)
    # Row index 2+ = data
    data_rows = rows[2:]

    # Build column→country_code mapping
    col_to_cc: dict[int, str] = {}
    for col_idx, header in enumerate(headers):
        if not header or col_idx == 0:
            continue
        cc = _COUNTRY_MAP.get(str(header).strip())
        if cc:
            col_to_cc[col_idx] = cc

    records: list[tuple[str, int, float]] = []
    for row in data_rows:
        if not row or row[0] is None:
            continue
        try:
            year = int(row[0])
        except (TypeError, ValueError):
            continue
        if year < 1800 or year > 2100:
            continue

        for col_idx, cc in col_to_cc.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None:
                continue
            try:
                records.append((cc, year, float(val)))
            except (TypeError, ValueError):
                continue

    return records


class WercScraper(BaseScraper):
    """
    Downloads the WERC Global Wine Markets megafile and imports vine area
    and wine production data into fact_werc_stats.
    """

    source_code = "werc"

    def __init__(self, conn: sqlite3.Connection):
        self.conn = conn
        self.batch_id: Optional[str] = None

    def _ensure_tables(self) -> None:
        self.conn.execute(_ENSURE_TABLE_SQL)
        self.conn.execute(_ENSURE_IDX_UNIQUE)
        self.conn.execute(_ENSURE_IDX_COUNTRY)
        self.conn.execute(_ENSURE_IDX_METRIC)
        self.conn.commit()

    def run(self, limit: Optional[int] = None) -> ScrapeResult:
        if not HAS_OPENPYXL:
            return ScrapeResult(
                batch_id="werc-no-openpyxl",
                error="openpyxl not installed. Run: pip install openpyxl",
            )

        self._ensure_tables()
        source_key = _get_source_key(self.conn, self.source_code)
        batch_id = self.batch_id or (
            f"werc-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}"
            f"-{uuid.uuid4().hex[:8]}"
        )
        result = ScrapeResult(batch_id=batch_id)

        # ── Download megafile ────────────────────────────────────────────────
        # Adelaide University's server uses legacy TLS renegotiation (disabled
        # in OpenSSL 3.x / Python 3.10+). We disable certificate verification
        # for this trusted academic host as the only practical workaround.
        console.print("[cyan]werc[/] downloading megafile (~17 MB)…")
        import ssl as _ssl
        # Adelaide University's server uses legacy TLS renegotiation, disabled
        # in OpenSSL 3.x. Build a permissive SSL context and pass it via verify=.
        ssl_ctx = _ssl.SSLContext(_ssl.PROTOCOL_TLS_CLIENT)
        ssl_ctx.check_hostname = False
        ssl_ctx.verify_mode = _ssl.CERT_NONE
        # OP_LEGACY_SERVER_CONNECT re-enables legacy renegotiation (Python 3.12+)
        if hasattr(_ssl, "OP_LEGACY_SERVER_CONNECT"):
            ssl_ctx.options |= _ssl.OP_LEGACY_SERVER_CONNECT

        with httpx.Client(timeout=120, follow_redirects=True, verify=ssl_ctx) as client:
            try:
                resp = self._fetch(lambda: client.get(_MEGAFILE_URL))
                resp.raise_for_status()
                xlsx_bytes = resp.content
            except Exception as exc:
                msg = f"Download failed: {exc}"
                console.print(f"[red]werc[/] {msg}")
                write_dlq(self.conn, source_key, batch_id, "parse_error", msg,
                          {"url": _MEGAFILE_URL})
                result.rows_dlq += 1
                result.error = msg
                return result

        console.print(f"[cyan]werc[/] {len(xlsx_bytes) / 1_048_576:.1f} MB downloaded — opening workbook…")

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(xlsx_bytes), read_only=True, data_only=True
            )
        except Exception as exc:
            result.error = f"openpyxl failed to open workbook: {exc}"
            return result

        # ── Parse each target sheet ─────────────────────────────────────────
        for sheet_frag, metric, unit in _SHEETS:
            ws = _find_sheet(wb, sheet_frag)
            if ws is None:
                console.print(f"[yellow]werc[/] sheet '{sheet_frag}' not found — skipping")
                continue

            records = _parse_sheet(ws, metric, unit)
            console.print(
                f"[cyan]werc[/] {sheet_frag}: {len(records)} country-year records parsed"
            )
            result.rows_fetched += len(records)

            for cc, year, value in records:
                if limit is not None and result.rows_inserted >= limit:
                    result.rows_skipped_unchanged += 1
                    continue
                try:
                    self.conn.execute(
                        """
                        INSERT OR REPLACE INTO fact_werc_stats
                            (source_key, country_code, year, metric, value, unit, batch_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (source_key, cc, year, metric, value, unit, batch_id),
                    )
                    result.rows_inserted += 1
                except Exception as exc:
                    write_dlq(self.conn, source_key, batch_id, "validation_error",
                              str(exc), {"cc": cc, "year": year, "metric": metric})
                    result.rows_dlq += 1

            self.conn.commit()

        wb.close()
        console.print(
            f"[green]werc[/] done — {result.rows_inserted} rows inserted, "
            f"{result.rows_dlq} DLQ"
        )
        return result
